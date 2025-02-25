
from os import close
from struct import pack
import pandas as pd
import pyautogui as ptg
import numpy
import customtkinter as ctk
import tkinter
from tkinter import messagebox
import openpyxl as op
from deep_translator import GoogleTranslator, single_detection


#tela de onde sera enviado a ultima msg do cliente para ser provado o tradutor com python na planilha

tela = ctk.CTk()
tela.geometry('450x450')
link = ctk.CTkEntry(tela, placeholder_text='link do chat', width=400)
link.pack(pady=10)

#email cliente
email = ctk.CTkEntry(tela, placeholder_text='email do cliente: ', width=400)
email.pack(pady=10)

#ultima msg do cliente

msg_ult = ctk.CTkEntry(tela, placeholder_text='ultima mensagem do cliente: ', width=400)
msg_ult.pack(pady=10)

#id usuario

id_user = ctk.CTkEntry(tela, placeholder_text='insira o id user: ', width=400)
id_user.pack(pady=10)
def savedados():
    tradutor = GoogleTranslator(source='auto', target='pt')
    traducao = tradutor.translate(msg_ult.get())
    language = single_detection(msg_ult.get(), api_key="8004ed31b3e275788568e01e161ae3fa")
    print(f'this languages is {language}')

    planilha = op.load_workbook("clientes.xlsx")# criar uma planilha para armazenar valores
    #nomeando as colunas
    msg = planilha['Planilha1']#nome da planilha
    msg = planilha.active
    msg['A1'] = "LINK_CHAT"
    msg['B1'] = "EMAIL_CLIENT"
    ulti_msg = msg['C1'] = 'MSG_ULT'
    msg['D1'] = 'ID_USER'
    trans = msg['E1'] = "TRANSL"
    idioma = msg['F1'] = 'LANGUAGE'  
    

    #planilha.save("clientes.xlsx") #salva  os dados da planilha

    #armazenando valores solicitado

    msg.cell(column=1, row=msg.max_row+1,  value=str(link.get())) #criar os index com ctk para armazenar as informações na planilha para poder validar a funcionalidade
    msg.cell(column=2, row=msg.max_row,  value=str(email.get()))
    msg.cell(column=3, row=msg.max_row,  value=str(msg_ult.get()))
    msg.cell(column=4, row=msg.max_row,  value=str(id_user.get()))
    msg.cell(column=5, row=msg.max_row,  value=str(traducao))
    msg.cell(column=6, row=msg.max_row,  value=str(language))
    #fazer a comparação de perguntas na planilha2 para retornar a resposta de acordo com a pergunta que mais se pareça com a do cliente 
    reply =  planilha['Planilha2']#nome da planilha
    pergunas_frequentes = reply["G1"] = 'PERGUNTAS_FRENQ'
    respostas = reply["H1"] = 'RESPOSTAS'
    if traducao in pergunas_frequentes:
        return respostas == pergunas_frequentes
    planilha.save("clientes.xlsx") #salva  os dados da planilha
    

btn0 = ctk.CTkButton(tela, text="enviar", command=savedados, width=250, height=35)
btn0.bind('<Return>', lambda e: btn0.configure(savedados()))
btn0.pack(pady=12)



tela.mainloop()

""""####salva a planilha de banco de dados e fazer a interface grafica hoje 03/08/2024####
tela = ctk.CTk()
email = ctk.CTkLabel(tela,text="Insira seu email.")
email.pack(pady=12)

email = ctk.CTkEntry(tela, width=400)
email.pack(pady=12)

url = ctk.CTkLabel(tela,text="Insira o Link do site")
url.pack(pady=12)
url = ctk.CTkEntry(tela,width=400)
url.pack(pady=12)


#saida de dados
perguntal = ctk.CTkLabel(tela,text="Qual e sua pergunta?")
perguntal.pack(pady=12)
pergunta = ctk.CTkEntry(tela, width=400)
pergunta.pack(pady=12)
def resposta():
    print(pergunta.get())
    




tela.geometry('450x450')

tela.mainloop()


planilha = op.load_workbook("planilha.xlsx")# criar uma planilha para armazenar valores
#nomeando as colunas
msg = planilha['']#nome da planilha
msg = planilha.active
msg['A1'] = "ULTIMA_MSG"
msg['B1'] = "LINK_CHAT"
msg['C1'] = 'ID_USER'
msg['D1'] = 'EMAIL'

planilha.save('planilha.xlsx') #salva  os dados da planilha

#armazenando valores solicitado

msg.cell(column=1, row=msg.max_row+1,  value=str("")) #criar os index com ctk para armazenar as informações na planilha para poder validar a funcionalidade
msg.cell(column=2, row=msg.max_row,  value=str(""))
msg.cell(column=3, row=msg.max_row,  value=str(""))
msg.cell(column=4, row=msg.max_row,  value=str(""))

planilha.save('nome da planilha')

"""
#quando for testar usar o tkinter msgbox para retornar as validações do sistema do bot, cada informação validada será retornada uma verificação com msgbox